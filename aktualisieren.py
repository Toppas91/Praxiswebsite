#!/usr/bin/env python3
# coding: utf-8
"""
aktualisieren.py  –  Praxis-Editor  (v2.1)
────────────────────────────────────────────
Erster Aufruf (Excel fehlt):
    Erstellt praxis-editor.xlsx aus praxis-daten.json.
    Bitte dann die Excel-Datei bearbeiten und das Script erneut starten.

Normaler Aufruf (Excel vorhanden):
    Liest praxis-editor.xlsx und schreibt Änderungen in praxis-daten.json.
    ► Schließzeiten (typ "schliessung") und Freitexte (typ "freitext")
      werden beide vollständig aus dem Sheet "Schließzeiten" gelesen.
    ► Das Sheet hat zwei Abschnitte: Schließzeiten (oben) + Meldungen (unten).

Aufruf mit --neu:
    Erstellt praxis-editor.xlsx neu aus der aktuellen praxis-daten.json
    (z. B. nach manueller JSON-Änderung, neuen Ärzten oder nach der
     einmaligen Migration von "schliesstage" → "events").

SHEET-LAYOUT "Schließzeiten":
    Zeile  1    : Titel
    Zeile  2    : Hinweis
    Zeile  3    : Gruppen-Überschriften (Schließzeit | Vertreter 1–4)
    Zeile  4    : Spalten-Überschriften
    Zeilen 5–19 : Schließzeiten-Daten (bis zu 15 Einträge)
    Zeile  20   : Trenner
    Zeile  21   : Abschnitts-Titel "AD-HOC-MELDUNGEN"
    Zeile  22   : Spalten-Überschriften (Von · Bis · Meldungstext)
    Zeilen 23–32: Freitext-Daten (bis zu 10 Einträge)
"""
import json
import sys
from datetime import datetime, time as dtime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("✗ openpyxl nicht gefunden.")
    print("  Bitte ausführen:  pip install openpyxl")
    sys.exit(1)

BASE        = Path(__file__).parent
JSON_FILE   = BASE / "praxis-daten.json"
EXCEL_FILE  = next(
    (BASE / n for n in ("praxis-editor.xlsm", "praxis-editor.xlsx")
     if (BASE / n).exists()),
    BASE / "praxis-editor.xlsm"
)
WOCHENTAGE  = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag"]


# ═══════════════════════════════════════════════════════════════
#  HILFSFUNKTIONEN
# ═══════════════════════════════════════════════════════════════

def als_zeit(val):
    """Zellwert → 'HH:MM' oder None."""
    if val is None:
        return None
    if isinstance(val, float):
        mins = round(val * 1440)
        return f"{mins // 60:02d}:{mins % 60:02d}"
    if isinstance(val, dtime):
        return val.strftime("%H:%M")
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    s = str(val).strip()
    if not s or s in ("-", "–", "—"):
        return None
    try:
        datetime.strptime(s, "%H:%M")
        return s
    except ValueError:
        return None

def als_datum(val):
    """Zellwert → 'YYYY-MM-DD' oder None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def als_int(val, default=7):
    if val is None:
        return default
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return default

def get_sheet(wb, token):
    """Findet ein Sheet anhand eines Teil-Namens (umlaut-tolerant)."""
    for name in wb.sheetnames:
        if token.lower() in name.lower():
            return wb[name]
    raise KeyError(f"Sheet mit '{token}' nicht gefunden. Vorhanden: {wb.sheetnames}")


# ═══════════════════════════════════════════════════════════════
#  JSON LESEN (EXCEL → JSON)
# ═══════════════════════════════════════════════════════════════

def lese_oeffnungszeiten(ws):
    tage = []
    for i, name in enumerate(WOCHENTAGE):
        r = 5 + i
        def z(col, _r=r): return als_zeit(ws.cell(row=_r, column=col).value)
        tage.append({
            "tag": name,
            "gesund": {"von": z(2), "bis": z(3), "nachmittag_von": z(4), "nachmittag_bis": z(5)},
            "krank":  {"von": z(6), "bis": z(7), "nachmittag_von": z(8), "nachmittag_bis": z(9)},
        })
    return tage


def _ist_neues_format(ws):
    """True, wenn das Sheet im neuen Format (v2.x) vorliegt.
    Erkennung: A3 = 'Schließzeit' (Gruppen-Überschrift) vs. 'Von' (Spalten-Überschrift)."""
    a3 = str(ws.cell(row=3, column=1).value or "").strip()
    return a3.lower() not in ("von", "")


def lese_schliesstage(ws, aerzte_pool):
    """
    Liest Schließzeiten (Zeilen 5–19) aus dem Schließzeiten-Sheet.
    Gibt Liste von Events mit typ="schliessung" zurück.
    Erkennt altes Format (v1.x, Daten ab Zeile 4) automatisch.
    """
    name_zu_id   = {a["name"]: a["id"] for a in aerzte_pool}
    neues_format = _ist_neues_format(ws)

    if not neues_format:
        print("  ℹ  Altes Excel-Format erkannt (keine Vertreter-/Freitext-Spalten).")
        print("     Tipp: 'python aktualisieren.py --neu' erstellt das neue Format.")
        start_row = 4
    else:
        start_row = 5

    ergebnis = []
    for row in range(start_row, start_row + 15):
        von     = als_datum(ws.cell(row=row, column=1).value)
        bis     = als_datum(ws.cell(row=row, column=2).value)
        grund   = str(ws.cell(row=row, column=3).value or "").strip()
        vorlauf = als_int(ws.cell(row=row, column=4).value, default=7)
        if not (von and bis):
            continue

        vertreter = []
        if neues_format:
            for v in range(4):
                base_col = 5 + v * 3          # Spalten 5, 8, 11, 14
                v_name   = str(ws.cell(row=row, column=base_col).value or "").strip()
                v_von    = als_datum(ws.cell(row=row, column=base_col + 1).value)
                v_bis    = als_datum(ws.cell(row=row, column=base_col + 2).value)
                if not v_name:
                    continue
                v_id = name_zu_id.get(v_name)
                if v_id is None:
                    print(f"  ⚠  Zeile {row}: Vertreter '{v_name}' nicht in der "
                          f"Vertretungsliste – übersprungen.")
                    continue
                entry = {"id": v_id}
                if v_von: entry["von"] = v_von
                if v_bis: entry["bis"] = v_bis
                vertreter.append(entry)

        ergebnis.append({
            "typ":          "schliessung",
            "von":          von,
            "bis":          bis,
            "grund":        grund,
            "vorlauf_tage": vorlauf,
            "vertreter":    vertreter,
        })
    return ergebnis


def lese_freitexte(ws):
    """
    Liest Ad-hoc-Meldungen (Zeilen 23–32) aus dem Schließzeiten-Sheet.
    Gibt Liste von Events mit typ="freitext" zurück.
    Leere Zeilen und Zeilen ohne Text werden übersprungen.
    Gibt leere Liste zurück, wenn altes Format erkannt wird.
    """
    if not _ist_neues_format(ws):
        return []

    ergebnis = []
    for row in range(23, 33):          # 10 Freitext-Slots, Zeilen 23–32
        von  = als_datum(ws.cell(row=row, column=1).value)
        bis  = als_datum(ws.cell(row=row, column=2).value)
        text = str(ws.cell(row=row, column=3).value or "").strip()
        if von and bis and text:
            ergebnis.append({
                "typ":  "freitext",
                "von":  von,
                "bis":  bis,
                "text": text,
            })
    return ergebnis


def lese_vertretung(ws, bestehende):
    id_map = {a["id"]: a for a in bestehende}
    neu = []
    for row in range(4, 30):
        id_val    = ws.cell(row=row, column=1).value
        aktiv_val = ws.cell(row=row, column=2).value
        if id_val is None:
            break
        arzt_id = str(id_val).strip()
        if arzt_id in id_map:
            arzt = dict(id_map[arzt_id])
            arzt["aktiv"] = str(aktiv_val or "").strip().upper() == "JA"
            neu.append(arzt)
    return neu or bestehende


def git_push():
    """Führt git add → commit → push aus. Gibt True bei Erfolg zurück."""
    import subprocess

    def git(args, **kw):
        return subprocess.run(
            ["git"] + args,
            cwd=BASE,
            capture_output=True,
            text=True,
            **kw
        )

    check = git(["rev-parse", "--is-inside-work-tree"])
    if check.returncode != 0:
        print("⚠ Kein Git-Repository gefunden – Push übersprungen.")
        return False

    status = git(["status", "--porcelain", str(JSON_FILE)])
    if not status.stdout.strip():
        print("  Keine Änderungen in praxis-daten.json – kein Commit nötig.")
        return True

    zeitstempel = datetime.now().strftime("%d.%m.%Y %H:%M")
    print("Git: Änderungen werden hochgeladen ...")
    schritte = [
        (["add", str(JSON_FILE)],                              "add"),
        (["commit", "-m", f"Praxisdaten aktualisiert ({zeitstempel})"], "commit"),
        (["push"],                                             "push"),
    ]

    for args, name in schritte:
        ergebnis = git(args)
        if ergebnis.returncode != 0:
            print(f"✗ Git-{name} fehlgeschlagen:")
            print(ergebnis.stderr.strip() or ergebnis.stdout.strip())
            print("  Bitte manuell pushen.")
            return False
        print(f"  ✓ git {name}")

    print("✓ Website wird in ~60 Sekunden aktuell sein.")
    return True


def aktualisiere():
    if not EXCEL_FILE.exists():
        print(f"✗ {EXCEL_FILE.name} nicht gefunden.")
        print("  Starte Script erneut ohne Argumente, um sie zu erstellen.")
        return False

    print(f"Lese {EXCEL_FILE.name} ...")
    wb = load_workbook(EXCEL_FILE, data_only=True)

    with open(JSON_FILE, encoding="utf-8") as f:
        daten = json.load(f)

    # ── Öffnungszeiten & Vertretungspool (unverändert) ───────────
    daten["oeffnungszeiten"]["tage"] = lese_oeffnungszeiten(
        get_sheet(wb, "ffnungszeiten")
    )
    daten["vertretung"]["aerzte"] = lese_vertretung(
        get_sheet(wb, "ertretung"), daten["vertretung"]["aerzte"]
    )

    # ── Events: beide Typen kommen jetzt aus demselben Sheet ────
    #    Schließzeiten aus Zeilen 5–19, Freitexte aus Zeilen 23–32.
    ws_schliess        = get_sheet(wb, "chlie")
    schliessung_events = lese_schliesstage(ws_schliess, daten["vertretung"]["aerzte"])
    freitext_events    = lese_freitexte(ws_schliess)
    daten["events"]    = schliessung_events + freitext_events

    # Legacy-Key entfernen (einmalige Migration)
    daten.pop("schliesstage", None)

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(daten, f, ensure_ascii=False, indent=2)

    print(f"✓ {JSON_FILE.name} erfolgreich aktualisiert.")
    print(f"  Schließzeiten: {len(schliessung_events)}"
          f"  |  Freitexte: {len(freitext_events)}")
    git_push()
    return True


# ═══════════════════════════════════════════════════════════════
#  EXCEL ERSTELLEN (JSON → EXCEL)
# ═══════════════════════════════════════════════════════════════

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

C  = Alignment(horizontal="center", vertical="center")
CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
L  = Alignment(horizontal="left",   vertical="center")

def titel_zeile(ws, text, ncols, bg):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill      = _fill(bg)
    c.alignment = C
    ws.row_dimensions[1].height = 30

def hinweis_zeile(ws, text, ncols):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=text)
    c.font      = _font(italic=True, size=9, color="888888")
    c.alignment = L
    ws.row_dimensions[2].height = 16

def kopf(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(bold=True)
    c.fill      = _fill("EEEEEE")
    c.alignment = CW
    c.border    = _border()
    return c

def eingabe(ws, row, col, val, fmt=None, readonly=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _font(size=10)
    c.fill      = _fill("F5F5F5") if readonly else _fill("FFFDE7")
    c.alignment = C
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    return c


def erstelle_excel(daten):
    # ── Bestehende xlsm mit VBA laden, falls vorhanden ──────────
    #    openpyxl speichert das VBA-Binary in wb.vba_archive.
    #    Sheets werden geleert und neu aufgebaut; der VBA-Modulcode
    #    (AppleScriptTask-Makro) bleibt vollständig erhalten.
    #    Einzige manuelle Nacharbeit: Schaltfläche im neuen Sheet
    #    neu zeichnen und dem vorhandenen Makronamen zuweisen.
    if EXCEL_FILE.exists() and EXCEL_FILE.suffix == ".xlsm":
        wb = load_workbook(EXCEL_FILE, keep_vba=True)
        for name in list(wb.sheetnames):
            del wb[name]
        print("  ℹ  Bestehendes xlsm geladen – VBA-Modul bleibt erhalten.")
        print("     Schaltfläche im Sheet 'Öffnungszeiten' neu anlegen")
        print("     und dem vorhandenen Makro zuweisen.")
    else:
        wb = Workbook()

    # ── Sheet 1: Öffnungszeiten ──────────────────────────────────
    # wb.active ist None, wenn alle Sheets geloescht wurden (xlsm-Pfad).
    # Deshalb immer explizit per create_sheet arbeiten.
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws1 = wb.create_sheet("Oeffnungszeiten")
    ws1.title = "Öffnungszeiten"

    titel_zeile(ws1, "ÖFFNUNGSZEITEN", 9, "2E7D32")
    hinweis_zeile(ws1,
        "Zeiten im Format HH:MM  ·  Kein Nachmittag? → Zelle leer lassen", 9)

    ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    c = ws1.cell(row=3, column=2, value="Gesundsprechstunde")
    c.font = _font(bold=True, color="1B5E20"); c.fill = _fill("C8E6C9")
    c.alignment = C; c.border = _border()

    ws1.merge_cells(start_row=3, start_column=6, end_row=3, end_column=9)
    c = ws1.cell(row=3, column=6, value="Kranksprechstunde")
    c.font = _font(bold=True, color="B71C1C"); c.fill = _fill("FFCDD2")
    c.alignment = C; c.border = _border()

    c = ws1.cell(row=3, column=1)
    c.fill = _fill("EEEEEE"); c.border = _border()
    ws1.row_dimensions[3].height = 20

    for col, h in enumerate(["Wochentag",
            "von", "bis", "Nachmittag von", "Nachmittag bis",
            "von", "bis", "Nachmittag von", "Nachmittag bis"], 1):
        kopf(ws1, 4, col, h)
    ws1.row_dimensions[4].height = 20

    tage = daten["oeffnungszeiten"]["tage"]
    for i, tag in enumerate(tage):
        row = 5 + i
        g, k = tag["gesund"], tag["krank"]
        c = ws1.cell(row=row, column=1, value=tag["tag"])
        c.font = _font(bold=True); c.fill = _fill("F5F5F5")
        c.alignment = C; c.border = _border()
        for col, val in enumerate([
            g.get("von") or "", g.get("bis") or "",
            g.get("nachmittag_von") or "", g.get("nachmittag_bis") or "",
            k.get("von") or "", k.get("bis") or "",
            k.get("nachmittag_von") or "", k.get("nachmittag_bis") or "",
        ], 2):
            eingabe(ws1, row, col, val, fmt="@")
        ws1.row_dimensions[row].height = 22

    ws1.column_dimensions["A"].width = 14
    for col in list("BCDEFGHI"):
        ws1.column_dimensions[col].width = 15
    ws1.freeze_panes = "B5"


    # ── Sheet 2: Schließzeiten (v2.0 – 16 Spalten mit Vertreter) ─
    ws2 = wb.create_sheet("Schließzeiten")
    NCOLS_S = 16

    titel_zeile(ws2, "SCHLIESS- & FERIENZEITEN", NCOLS_S, "C62828")
    hinweis_zeile(ws2,
        "OBEN – Schließzeiten: Datum TT.MM.JJJJ · Vorlauf = Tage bis Banner erscheint · "
        "Vertreter: Name aus Dropdown, Von/Bis leer = gesamte Schließzeit  ·  "
        "UNTEN – Ad-hoc-Meldungen: Von/Bis + Freitext → erscheint sofort im Banner",
        NCOLS_S)
    ws2.row_dimensions[2].height = 18

    # Zeile 3: Gruppen-Überschriften
    ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    c = ws2.cell(row=3, column=1, value="Schließzeit")
    c.font = _font(bold=True, color="B71C1C"); c.fill = _fill("FFCDD2")
    c.alignment = C; c.border = _border()

    vertreter_bg   = ["D6E4F7", "D4EDD4", "FEF3CD", "EDE0F5"]
    vertreter_text = ["1565C0", "2E7D32", "856404", "6A1B9A"]
    for v_idx, v_start in enumerate([5, 8, 11, 14]):
        ws2.merge_cells(
            start_row=3, start_column=v_start,
            end_row=3,   end_column=v_start + 2
        )
        c = ws2.cell(row=3, column=v_start, value=f"Vertreter {v_idx + 1}")
        c.font      = _font(bold=True, color=vertreter_text[v_idx])
        c.fill      = _fill(vertreter_bg[v_idx])
        c.alignment = C; c.border = _border()
    ws2.row_dimensions[3].height = 18

    # Zeile 4: Spalten-Überschriften
    for col, h in enumerate(["Von", "Bis", "Grund", "Vorlauf\n(Tage)"], 1):
        kopf(ws2, 4, col, h)
    for v_idx in range(4):
        base = 5 + v_idx * 3
        kopf(ws2, 4, base,     "Name")
        kopf(ws2, 4, base + 1, "Von\n(opt.)")
        kopf(ws2, 4, base + 2, "Bis\n(opt.)")
    ws2.row_dimensions[4].height = 28

    # Dropdown-Validierung: Arzt-Namen aus dem Vertretungspool
    aerzte_pool = daten["vertretung"]["aerzte"]
    id_zu_name  = {a["id"]: a["name"] for a in aerzte_pool}
    namen_csv   = ",".join(a["name"] for a in aerzte_pool)

    dv_vertreter = DataValidation(
        type="list",
        formula1=f'"{namen_csv}"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv_vertreter.error      = "Bitte einen Namen aus dem Dropdown wählen oder Zelle leer lassen."
    dv_vertreter.errorTitle = "Ungültige Eingabe"
    ws2.add_data_validation(dv_vertreter)

    # Bestehende Daten aus JSON laden (Backward-Compat: auch altes schliesstage-Format)
    events = daten.get("events", [])
    if not any(e.get("typ") == "schliessung" for e in events) and "schliesstage" in daten:
        # Einmalige Migration: altes Format in events konvertieren
        events = events + [
            {"typ": "schliessung", "von": s["von"], "bis": s["bis"],
             "grund": s.get("grund", ""), "vorlauf_tage": s.get("vorlauf_tage", 7),
             "vertreter": []}
            for s in daten["schliesstage"]
        ]
    schliessung_events = [e for e in events if e.get("typ") == "schliessung"]

    # Datenzeilen 5–19
    for i in range(15):
        row = 5 + i

        if i < len(schliessung_events):
            ev = schliessung_events[i]
            try:
                von_dt = datetime.strptime(ev["von"], "%Y-%m-%d")
                bis_dt = datetime.strptime(ev["bis"], "%Y-%m-%d")
            except Exception:
                von_dt, bis_dt = "", ""
            basis_vals = [von_dt, bis_dt, ev.get("grund", ""), ev.get("vorlauf_tage", 7)]
            basis_fmts = ["DD.MM.YYYY", "DD.MM.YYYY", "@", "0"]
            vertreter_liste = ev.get("vertreter", [])
        else:
            basis_vals      = ["", "", "", ""]
            basis_fmts      = ["DD.MM.YYYY", "DD.MM.YYYY", "@", "0"]
            vertreter_liste = []

        # Schließzeit-Spalten (A–D)
        for col, (val, fmt) in enumerate(zip(basis_vals, basis_fmts), 1):
            eingabe(ws2, row, col, val, fmt=fmt)

        # Vertreter-Spalten (E–P, je 3 Spalten pro Vertreter)
        for v_idx in range(4):
            base_col = 5 + v_idx * 3
            if v_idx < len(vertreter_liste):
                vr     = vertreter_liste[v_idx]
                v_name = id_zu_name.get(vr["id"], vr["id"])
                try:
                    v_von = datetime.strptime(vr["von"], "%Y-%m-%d") if vr.get("von") else ""
                    v_bis = datetime.strptime(vr["bis"], "%Y-%m-%d") if vr.get("bis") else ""
                except Exception:
                    v_von, v_bis = "", ""
            else:
                v_name, v_von, v_bis = "", "", ""

            c_name = eingabe(ws2, row, base_col,     v_name, fmt="@")
            dv_vertreter.add(c_name)
            eingabe(ws2, row, base_col + 1, v_von, fmt="DD.MM.YYYY")
            eingabe(ws2, row, base_col + 2, v_bis, fmt="DD.MM.YYYY")

        ws2.row_dimensions[row].height = 22

    # Spaltenbreiten
    breiten = {
        "A": 16, "B": 16, "C": 34, "D": 14,
        "E": 30, "F": 14, "G": 14,
        "H": 30, "I": 14, "J": 14,
        "K": 30, "L": 14, "M": 14,
        "N": 30, "O": 14, "P": 14,
    }
    for col_letter, width in breiten.items():
        ws2.column_dimensions[col_letter].width = width

    # ── Abschnitt 2: Ad-hoc-Meldungen (Freitext), Zeilen 20–32 ──

    # Zeile 20: visueller Trenner
    ws2.merge_cells(start_row=20, start_column=1, end_row=20, end_column=NCOLS_S)
    ws2.cell(row=20, column=1).fill = _fill("EFEFEF")
    ws2.row_dimensions[20].height = 10

    # Zeile 21: Abschnitts-Titel
    ws2.merge_cells(start_row=21, start_column=1, end_row=21, end_column=NCOLS_S)
    c = ws2.cell(row=21, column=1, value="AD-HOC-MELDUNGEN  ·  FREITEXT-BANNER")
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = _fill("E65100")   # kräftiges Orange – unterscheidet sich visuell von Rot oben
    c.alignment = C
    ws2.row_dimensions[21].height = 26

    # Zeile 22: Spalten-Überschriften
    kopf(ws2, 22, 1, "Von")
    kopf(ws2, 22, 2, "Bis")
    ws2.merge_cells(start_row=22, start_column=3, end_row=22, end_column=NCOLS_S)
    c = ws2.cell(row=22, column=3, value="Meldungstext (erscheint so im Banner)")
    c.font = _font(bold=True); c.fill = _fill("EEEEEE")
    c.alignment = CW; c.border = _border()
    ws2.row_dimensions[22].height = 20

    # Datenzeilen 23–32 (10 Slots)
    freitext_events = [e for e in events if e.get("typ") == "freitext"]
    for i in range(10):
        row = 23 + i
        if i < len(freitext_events):
            fe = freitext_events[i]
            try:
                ft_von = datetime.strptime(fe["von"], "%Y-%m-%d")
                ft_bis = datetime.strptime(fe["bis"], "%Y-%m-%d")
            except Exception:
                ft_von, ft_bis = "", ""
            ft_text = fe.get("text", "")
        else:
            ft_von, ft_bis, ft_text = "", "", ""

        eingabe(ws2, row, 1, ft_von, fmt="DD.MM.YYYY")
        eingabe(ws2, row, 2, ft_bis, fmt="DD.MM.YYYY")

        ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=NCOLS_S)
        c = ws2.cell(row=row, column=3, value=ft_text)
        c.font      = _font(size=10)
        c.fill      = _fill("FFFDE7")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = _border()
        ws2.row_dimensions[row].height = 36

    ws2.freeze_panes = "A5"


    # ── Sheet 3: Vertretung ──────────────────────────────────────
    ws3 = wb.create_sheet("Vertretung")
    titel_zeile(ws3, "VERTRETUNGSÄRZTE", 6, "1565C0")
    hinweis_zeile(ws3,
        'Aktiv → "JA": Arzt erscheint auf der Website  ·  '
        '"NEIN": ausgeblendet  ·  Alle anderen Felder nur zur Ansicht',
        6)

    for col, h in enumerate(["ID", "Aktiv", "Name", "Fachrichtung", "Adresse", "Telefon"], 1):
        kopf(ws3, 3, col, h)
    ws3.row_dimensions[3].height = 20

    dv = DataValidation(type="list", formula1='"JA,NEIN"', allow_blank=False)
    dv.error      = 'Bitte "JA" oder "NEIN" wählen.'
    dv.errorTitle = "Ungültige Eingabe"
    ws3.add_data_validation(dv)

    for i, arzt in enumerate(aerzte_pool):
        row   = 4 + i
        aktiv = "JA" if arzt.get("aktiv", True) else "NEIN"

        c = ws3.cell(row=row, column=1, value=arzt["id"])
        c.font = _font(size=9, color="AAAAAA")
        c.fill = _fill("F0F0F0"); c.border = _border()

        c = ws3.cell(row=row, column=2, value=aktiv)
        c.font      = _font(bold=True, color="1B5E20" if aktiv == "JA" else "B71C1C")
        c.fill      = _fill("E8F5E9") if aktiv == "JA" else _fill("FFEBEE")
        c.alignment = C; c.border = _border()
        dv.add(c)

        for col, val in enumerate([
            arzt["name"], arzt.get("fachrichtung", ""),
            arzt.get("adresse", ""), arzt.get("telefon", "")
        ], 3):
            eingabe(ws3, row, col, val, readonly=True).alignment = L
        ws3.row_dimensions[row].height = 22

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 28
    ws3.column_dimensions["D"].width = 36
    ws3.column_dimensions["E"].width = 40
    ws3.column_dimensions["F"].width = 18
    ws3.column_dimensions["A"].hidden = True
    ws3.freeze_panes = "B4"

    wb.save(EXCEL_FILE)
    print(f"✓ {EXCEL_FILE.name} wurde erstellt.")


# ═══════════════════════════════════════════════════════════════
#  EINSTIEGSPUNKT
# ═══════════════════════════════════════════════════════════════

def main():
    neu_erstellen = "--neu" in sys.argv

    with open(JSON_FILE, encoding="utf-8") as f:
        daten = json.load(f)

    if neu_erstellen:
        print("Erstelle praxis-editor.xlsx neu aus praxis-daten.json ...")
        erstelle_excel(daten)
        print("Fertig. Bitte Excel öffnen, prüfen und danach Script normal starten.")
        return True

    if not EXCEL_FILE.exists():
        print(f"{EXCEL_FILE.name} nicht gefunden – wird jetzt erstellt ...")
        erstelle_excel(daten)
        print()
        print("Bitte praxis-editor.xlsx jetzt bearbeiten und das Script")
        print("danach erneut starten, um die JSON zu aktualisieren.")
        return True

    return aktualisiere()


if __name__ == "__main__":
    try:
        ok = main()
    except Exception as e:
        print(f"✗ Fehler: {e}")
        ok = False
    sys.exit(0 if ok else 1)
